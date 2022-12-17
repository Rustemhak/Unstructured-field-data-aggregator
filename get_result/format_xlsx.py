from openpyxl import Workbook, load_workbook
import datetime

def merge_cells(path):
    #path = 'archangelsk.xlsx'
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    # print(ws['B2'].value)
    # print(ws['B3'].value)
    start_col = 'B'
    start_row = 2
    cur_col = start_col
    cur_row = start_row + 1
    while ws[cur_col + str(cur_row)].value is None:
        cur_col = chr(ord(cur_col) + 1)

    while ws[cur_col + str(cur_row)].value is not None:
        cur_row += 1
        print(ws[cur_col + str(cur_row)].value)
    end_col = chr(ord(cur_col) - 1)
    end_row = cur_row - 1
    # start_cell = start_col + str(start_row)
    # end_cell = end_col + str(end_row)
    # ws.merge_cells(f'{start_cell}:{end_cell}')
    # wb.save("styled_megre.xlsx")
    cur_col = start_col
    while cur_col <= end_col:
        start_cell = cur_col + str(start_row)
        end_cell = cur_col + str(end_row)
        ws.merge_cells(f'{start_cell}:{end_cell}')
        cur_col = chr(ord(cur_col) + 1)
    wb.save("styled_megre.xlsx")
